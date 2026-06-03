import os
import re
import io
import zipfile
from copy import copy
from datetime import datetime, date, time
from collections import defaultdict

import streamlit as st
import fitz  # PyMuPDF
import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
from pypdf import PdfReader, PdfWriter


# ============================================================
# SETTINGS
# ============================================================

OUTPUT_FILE_NAME = "Populated_KS_SS_TEMPLATE.xlsx"

# ============================================================
# BUNDLED DISPATCH TEMPLATE
# The template is a normal Excel file kept in the same folder as
# this script, so you can open and edit it directly in Excel.
# The loader accepts the file whether it is named with a space
# ("Dispatch Template.xlsx") or an underscore ("Dispatch_Template.xlsx").
# ============================================================

TEMPLATE_FILENAME = "Dispatch_Template.xlsx"


def _base_dirs() -> list[str]:
    dirs = []
    try:
        dirs.append(os.path.dirname(os.path.abspath(__file__)))
    except NameError:
        pass
    dirs.append(os.getcwd())
    seen = set()
    return [d for d in dirs if d and not (d in seen or seen.add(d))]


def _candidate_filenames() -> list[str]:
    names = [
        TEMPLATE_FILENAME,
        TEMPLATE_FILENAME.replace("_", " "),
        TEMPLATE_FILENAME.replace(" ", "_"),
    ]
    seen = set()
    return [n for n in names if not (n in seen or seen.add(n))]


def get_template_path() -> str:
    for base_dir in _base_dirs():
        for name in _candidate_filenames():
            candidate = os.path.join(base_dir, name)
            if os.path.exists(candidate):
                return candidate
    # Nothing found; return the primary expected path for the error message.
    return os.path.join(_base_dirs()[0], TEMPLATE_FILENAME)


def get_embedded_template_bytes() -> bytes:
    path = get_template_path()
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Dispatch template not found: {path}. "
            f"Place '{TEMPLATE_FILENAME}' (or 'Dispatch Template.xlsx') "
            f"in the same folder as this app."
        )
    with open(path, "rb") as f:
        return f.read()

DISPATCH_SHEET = "DISPATCH SHEET"
OPENDOCK_SHEET = "OPENDOCK"
MG_REPORT_SHEET = "MG REPORT"

TEMPLATE_REQUIRED_SHEETS = [
    DISPATCH_SHEET,
    OPENDOCK_SHEET,
    MG_REPORT_SHEET,
]

DISPATCH_START_ROW = 3
OPENDOCK_START_ROW = 2
MG_REPORT_START_ROW = 2

MAX_DISPATCH_ROWS = 300
MAX_OPENDOCK_ROWS = 1000
MAX_MG_REPORT_ROWS = 1000


# ============================================================
# DISPATCH SHEET COLUMN MAP
# ------------------------------------------------------------
# Columns are resolved at runtime from the header row (row 2) by their
# header text, NOT by fixed positions. This means you can rearrange,
# hide, or shift columns in Dispatch_Template.xlsx in Excel and the app
# keeps working as long as the header names below still exist.
#
# Header text on row 2 -> field (accepted aliases in parentheses):
#   LOAD#  (TRIP#)        -> load number               (written directly)
#   CUSTOMER (DESTINATION)-> consignee                 (VLOOKUP MG col 4)
#   CARRIER               -> carrier short code        (VLOOKUP on CARRIER FULL,
#                                                        with CPU fallback)
#   CARRIER FULL          -> full carrier name         (written directly)
#   TYPE                  -> cleaned load type         (written directly)
#   TIME   (DISPATCH)     -> appointment time          (written directly)
#   TT4                   -> flag                       (SEARCH formula on CUSTOMER)
#   WEIGHT                -> weight                      (VLOOKUP MG col 2)
#   CASES                 -> cases                       (VLOOKUP MG col 3)
#   PULLS / PICKS         -> kept as template formulas   (number format only)
#   NOTES  (NOTE)         -> match notes                 (written directly)
#
# Row 1 summary cells (load count + date) are also detected automatically.
# ============================================================

# How each logical field maps to acceptable header texts on the header row.
DISPATCH_HEADER_ALIASES = {
    "load":         ["LOAD#", "LOAD", "TRIP#", "TRIP"],
    "customer":     ["CUSTOMER", "DESTINATION"],
    "carrier":      ["CARRIER"],
    "carrier_full": ["CARRIER FULL"],
    "type":         ["TYPE", "TYPE / TR", "TYPE/TR"],
    "time":         ["TIME", "DISPATCH"],
    "tt4":          ["TT4"],
    "weight":       ["WEIGHT"],
    "cases":        ["CASES"],
    "pulls":        ["PULLS"],
    "picks":        ["PICKS"],
    "notes":        ["NOTES", "NOTE"],
}

# Fields the app must be able to find, or it cannot fill the sheet.
DISPATCH_REQUIRED_FIELDS = [
    "load", "customer", "carrier", "carrier_full",
    "type", "time", "tt4", "weight", "cases", "notes",
]


def get_dispatch_columns(ws, header_row=2):
    """Resolve {field: column_index} from the DISPATCH SHEET header row."""
    header_map = get_header_map(ws, header_row)

    columns = {}
    for field, aliases in DISPATCH_HEADER_ALIASES.items():
        columns[field] = find_col(header_map, aliases)

    missing = [
        field for field in DISPATCH_REQUIRED_FIELDS
        if not columns.get(field)
    ]
    if missing:
        found = ", ".join(
            f"{get_column_letter(c)}={ws.cell(header_row, c).value!r}"
            for c in range(1, ws.max_column + 1)
            if ws.cell(header_row, c).value not in (None, "")
        )
        raise ValueError(
            "DISPATCH SHEET is missing required column header(s): "
            + ", ".join(m.upper() for m in missing)
            + f". Headers found on row {header_row}: {found}"
        )

    return columns


def find_dispatch_count_cell(ws, header_row=1):
    """Find the row-1 cell holding the =COUNT(...) load-count formula."""
    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if isinstance(value, str) and "COUNT(" in value.upper():
            return ws.cell(header_row, col).coordinate
    return None


def find_dispatch_date_cell(ws, header_row=1):
    """Find the cell just right of the 'Date:' label on row 1."""
    for col in range(1, ws.max_column + 1):
        if normalize_header(ws.cell(header_row, col).value) == "DATE":
            return ws.cell(header_row, col + 1).coordinate
    return None


# ============================================================
# BASIC HELPERS
# ============================================================

def clean_spaces(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def clean_load_type(value) -> str:
    text = clean_spaces(value)

    text = re.sub(r"\bTRAILER\b", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\bLOAD\b", "", text, flags=re.IGNORECASE)

    return clean_spaces(text)

def normalize_header(value) -> str:
    if value is None:
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(value).strip().upper())


def normalize_load_number(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\.0$", "", text)
    text = re.sub(r"^LD", "", text, flags=re.IGNORECASE)
    text = re.sub(r"[^0-9]", "", text)
    return text if text.isdigit() else ""


def parse_number(value):
    if value is None:
        return 0
    text = str(value).replace(",", "").strip()
    if not text:
        return 0
    try:
        number = float(text)
        if number.is_integer():
            return int(number)
        return number
    except Exception:
        return 0


def normalize_date(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y")
    if isinstance(value, date):
        return value.strftime("%m/%d/%Y")
    text = clean_spaces(value)
    if not text:
        return ""
    match = re.search(r"(\d{1,2}/\d{1,2}/\d{2,4})", text)
    if match:
        text = match.group(1)
    for fmt in ["%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"]:
        try:
            dt = datetime.strptime(text, fmt)
            return dt.strftime("%m/%d/%Y")
        except Exception:
            pass
    return ""


def normalize_time(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, (int, float)) and 0 <= value < 1:
        total_minutes = int(round(value * 24 * 60))
        hours = (total_minutes // 60) % 24
        minutes = total_minutes % 60
        return f"{hours:02d}:{minutes:02d}"
    text = clean_spaces(value)
    if not text:
        return ""
    match = re.search(r"\b(\d{1,2}):(\d{2})\b", text)
    if match and "AM" not in text.upper() and "PM" not in text.upper():
        return f"{int(match.group(1)):02d}:{match.group(2)}"
    compact_text = text.upper().replace(" ", "")
    try:
        dt = datetime.strptime(compact_text, "%I:%M%p")
        return dt.strftime("%H:%M")
    except Exception:
        pass
    return ""


def sort_datetime_key(record):
    date_text = record.get("appt_date", "") or record.get("arrival_date", "")
    time_text = record.get("appt_time", "") or record.get("arrival_time", "")
    if not date_text:
        return datetime.max
    if not time_text:
        time_text = "23:59"
    try:
        return datetime.strptime(f"{date_text} {time_text}", "%m/%d/%Y %H:%M")
    except Exception:
        return datetime.max


def get_header_map(ws, header_row=1):
    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if value not in [None, ""]:
            headers[normalize_header(value)] = col
    return headers


def find_col(header_map, possible_names):
    for name in possible_names:
        key = normalize_header(name)
        if key in header_map:
            return header_map[key]
    return None


def find_dispatch_notes_col(ws, header_row=2):
    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if normalize_header(value) in ["NOTES", "NOTE"]:
            return col
    return None


def copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)


def copy_row_format(ws, source_row, target_row, max_col):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, max_col + 1):
        source_cell = ws.cell(source_row, col)
        target_cell = ws.cell(target_row, col)
        copy_cell_style(source_cell, target_cell)
        if isinstance(source_cell.value, str) and source_cell.value.startswith("="):
            target_cell.value = Translator(
                source_cell.value,
                origin=source_cell.coordinate
            ).translate_formula(target_cell.coordinate)


def clear_range_values(ws, start_row, end_row, columns):
    for row in range(start_row, end_row + 1):
        for col in columns:
            ws.cell(row, col).value = None


def capture_row_template(ws, source_row, max_col):
    template = {
        "height": ws.row_dimensions[source_row].height,
        "cells": []
    }

    for col in range(1, max_col + 1):
        cell = ws.cell(source_row, col)
        template["cells"].append({
            "col": col,
            "coordinate": cell.coordinate,
            "value": cell.value,
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "number_format": cell.number_format,
            "protection": copy(cell.protection),
        })

    return template


def apply_row_template(ws, row_template, target_row, copy_values=True):
    ws.row_dimensions[target_row].height = row_template["height"]

    for item in row_template["cells"]:
        col = item["col"]
        target_cell = ws.cell(target_row, col)

        target_cell.font = copy(item["font"])
        target_cell.fill = copy(item["fill"])
        target_cell.border = copy(item["border"])
        target_cell.alignment = copy(item["alignment"])
        target_cell.number_format = item["number_format"]
        target_cell.protection = copy(item["protection"])

        if copy_values:
            value = item["value"]

            if isinstance(value, str) and value.startswith("="):
                target_cell.value = Translator(
                    value,
                    origin=item["coordinate"]
                ).translate_formula(target_cell.coordinate)
            else:
                target_cell.value = value


def clear_dispatch_area(ws, start_row, end_row, max_col):
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).value = None


def write_day_header(ws, row, date_text, load_count, total_cases, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row, col)
        cell.value = None
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row, 2).value = f"{date_text} | LOADS: {load_count} | CASES: {int(round(total_cases, 0))}"
    ws.row_dimensions[row].height = 22


def extract_pdf_bytes_from_upload(uploaded_file) -> bytes:
    raw = uploaded_file.read()

    if uploaded_file.name.lower().endswith(".pdf"):
        return raw

    with zipfile.ZipFile(io.BytesIO(raw)) as z:
        pdf_names = sorted(
            name for name in z.namelist()
            if name.lower().endswith(".pdf")
            and not os.path.basename(name).startswith("__MACOSX")
            and not os.path.basename(name).startswith(".")
        )

        if not pdf_names:
            raise ValueError(f"No PDF files found inside {uploaded_file.name}.")

        if len(pdf_names) == 1:
            return z.read(pdf_names[0])

        writer = PdfWriter()

        for name in pdf_names:
            reader = PdfReader(io.BytesIO(z.read(name)))
            for page in reader.pages:
                writer.add_page(page)

        out = io.BytesIO()
        writer.write(out)
        out.seek(0)

        return out.read()


# ============================================================
# PDF MANIFEST MATCHER
# ============================================================

def sample_text_from_bytes(pdf_bytes, pages=5):
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    text = ""

    for i in range(min(pages, len(doc))):
        text += doc[i].get_text("text") + "\n"

    doc.close()

    return text.upper()


def parse_load_from_text(text):
    m = re.search(r"\bLOAD\s*:\s*([A-Z]{1,5}\d+)", text, re.I)
    return m.group(1).upper().strip() if m else None


def parse_pu_appt_from_text(text):
    m = re.search(
        r"\bPU\s+APPT\s*:\s*(\d{1,2}/\d{1,2}/\d{4})\s+(\d{1,2}:\d{2})",
        text,
        re.I
    )

    if not m:
        return None

    try:
        return datetime.strptime(m.group(1) + " " + m.group(2), "%m/%d/%Y %H:%M")
    except Exception:
        return None


def group_pages_by_load_from_bytes(pdf_bytes):
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    groups = {}
    current_load = None

    for i in range(len(doc)):
        text = doc[i].get_text("text")
        found_load = parse_load_from_text(text)

        if found_load:
            current_load = found_load

            if current_load not in groups:
                groups[current_load] = {
                    "pages": [],
                    "text": ""
                }

        if current_load:
            groups[current_load]["pages"].append(i)
            groups[current_load]["text"] += "\n" + text

    doc.close()

    return groups


def build_matched_pdf_bytes(loading_bytes, shipping_bytes):
    loading_text_sample = sample_text_from_bytes(loading_bytes)
    shipping_text_sample = sample_text_from_bytes(shipping_bytes)

    if "LOADING MANIFEST" in loading_text_sample and "SHIPPING MANIFEST" not in loading_text_sample:
        pass
    elif "LOADING MANIFEST" in shipping_text_sample and "SHIPPING MANIFEST" not in shipping_text_sample:
        loading_bytes, shipping_bytes = shipping_bytes, loading_bytes
    else:
        pass

    loading_groups = group_pages_by_load_from_bytes(loading_bytes)
    shipping_groups = group_pages_by_load_from_bytes(shipping_bytes)

    all_loads = sorted(
        set(loading_groups.keys()) |
        set(shipping_groups.keys())
    )

    records = []

    for load in all_loads:
        lt = loading_groups.get(load, {}).get("text", "")
        st_text = shipping_groups.get(load, {}).get("text", "")
        dt = parse_pu_appt_from_text(lt or st_text)

        records.append({
            "load": load,
            "datetime": dt,
            "loading_pages": loading_groups.get(load, {}).get("pages", []),
            "shipping_pages": shipping_groups.get(load, {}).get("pages", []),
        })

    records = sorted(
        records,
        key=lambda r: (
            r["datetime"] is None,
            r["datetime"] or datetime.max,
            r["load"],
        )
    )

    writer = PdfWriter()
    loading_reader = PdfReader(io.BytesIO(loading_bytes))
    shipping_reader = PdfReader(io.BytesIO(shipping_bytes))

    for r in records:
        for page_num in r["loading_pages"]:
            writer.add_page(loading_reader.pages[page_num])

        for page_num in r["shipping_pages"]:
            writer.add_page(shipping_reader.pages[page_num])

    output = io.BytesIO()
    writer.write(output)
    output.seek(0)

    return output.read(), len(loading_groups), len(shipping_groups), len(records)


# ============================================================
# OPENDOCK PARSER
# ============================================================

def parse_opendock_excel(opendock_bytes: bytes) -> tuple[list[dict], dict]:
    wb = openpyxl.load_workbook(io.BytesIO(opendock_bytes), data_only=True)

    if "Appointments" in wb.sheetnames:
        ws = wb["Appointments"]
    else:
        ws = wb[wb.sheetnames[0]]

    headers = get_header_map(ws, 1)

    load_col = find_col(headers, ["Load Reference", "Load", "Load #"])
    appt_date_col = find_col(headers, ["Appt Date", "Appointment Date", "Date"])
    appt_time_col = find_col(headers, ["Appt Time", "Appointment Time", "Time"])
    arrival_date_col = find_col(headers, ["Arrival Date"])
    arrival_time_col = find_col(headers, ["Arrival Time"])
    departure_date_col = find_col(headers, ["Departure Date"])
    departure_time_col = find_col(headers, ["Departure Time"])
    status_col = find_col(headers, ["Status"])
    carrier_col = find_col(headers, ["Carrier Company", "Carrier"])
    load_type_col = find_col(headers, ["Load Type", "Type"])
    dock_col = find_col(headers, ["Dock"])
    direction_col = find_col(headers, ["Direction"])

    missing = []

    if not load_col:
        missing.append("Load Reference")

    if not appt_date_col:
        missing.append("Appt Date")

    if not appt_time_col:
        missing.append("Appt Time")

    if not carrier_col:
        missing.append("Carrier Company")

    if missing:
        raise ValueError("The Opendock report is missing these columns: " + ", ".join(missing))

    records = []
    skipped_inbound = []
    duplicate_tracker = defaultdict(int)

    for row in range(2, ws.max_row + 1):
        load_number = normalize_load_number(ws.cell(row, load_col).value)

        if not load_number:
            continue

        direction = clean_spaces(ws.cell(row, direction_col).value) if direction_col else ""

        if direction and direction.upper() != "OUTBOUND":
            skipped_inbound.append(load_number)
            continue

        appt_date = normalize_date(ws.cell(row, appt_date_col).value)
        appt_time = normalize_time(ws.cell(row, appt_time_col).value)

        duplicate_tracker[load_number] += 1

        record = {
            "load": load_number,
            "appt_date": appt_date,
            "appt_time": appt_time,
            "arrival_date": normalize_date(ws.cell(row, arrival_date_col).value) if arrival_date_col else "",
            "arrival_time": normalize_time(ws.cell(row, arrival_time_col).value) if arrival_time_col else "",
            "departure_date": normalize_date(ws.cell(row, departure_date_col).value) if departure_date_col else "",
            "departure_time": normalize_time(ws.cell(row, departure_time_col).value) if departure_time_col else "",
            "status": clean_spaces(ws.cell(row, status_col).value) if status_col else "",
            "carrier": clean_spaces(ws.cell(row, carrier_col).value) if carrier_col else "",
            "load_type": clean_spaces(ws.cell(row, load_type_col).value) if load_type_col else "",
            "dock": clean_spaces(ws.cell(row, dock_col).value) if dock_col else "",
            "direction": direction,
            "source_row": row,
        }

        records.append(record)

    records.sort(key=lambda x: (sort_datetime_key(x), x.get("load", "")))

    duplicates = [
        load for load, count in duplicate_tracker.items()
        if count > 1
    ]

    summary = {
        "outbound_rows_loaded": len(records),
        "inbound_rows_skipped": len(skipped_inbound),
        "duplicate_outbound_loads": duplicates,
    }

    return records, summary


# ============================================================
# MG EXCEL PARSER
# ============================================================

def parse_mg_report_excel(mg_report_bytes: bytes, file_name: str = "") -> tuple[list[dict], dict]:
    if file_name.lower().endswith(".xls"):
        df = pd.read_excel(io.BytesIO(mg_report_bytes), engine="xlrd")
    else:
        df = pd.read_excel(io.BytesIO(mg_report_bytes), engine="openpyxl")

    df = df.fillna("")

    normalized_headers = {
        normalize_header(col): col
        for col in df.columns
    }

    def find_df_col(possible_names):
        for name in possible_names:
            key = normalize_header(name)
            if key in normalized_headers:
                return normalized_headers[key]
        return None

    load_col = find_df_col(["Load", "Load #", "Load Number", "Load Reference"])
    date_col = find_df_col(["PU Appt Date", "Appt Date", "Appointment Date", "Date"])
    time_col = find_df_col(["PU Appt Time", "Appt Time", "Appointment Time", "Time"])
    carrier_col = find_df_col(["Carrier", "Carrier Name", "CARR/SCT TR"])
    weight_col = find_df_col(["Weight", "Actual Weight", "Total Weight"])
    cases_col = find_df_col(["Quantity", "Actual Quantity", "Cases", "Case Count", "Total Cases"])
    customer_col = find_df_col(["Consignee Name", "Customer", "Customer Name", "Ship To", "Destination"])

    missing = []

    if not load_col:
        missing.append("Load")
    if not weight_col:
        missing.append("Weight")
    if not cases_col:
        missing.append("Quantity / Cases")
    if not customer_col:
        missing.append("Customer / Consignee")

    if missing:
        raise ValueError("The MG report Excel is missing these columns: " + ", ".join(missing))

    all_records = []
    duplicate_mg_loads = []
    records_by_load = {}

    for _, row in df.iterrows():
        load_number = normalize_load_number(row.get(load_col, ""))

        if not load_number:
            continue

        record = {
            "load": load_number,
            "appt_date": normalize_date(row.get(date_col, "")) if date_col else "",
            "appt_time": normalize_time(row.get(time_col, "")) if time_col else "",
            "carrier": clean_spaces(row.get(carrier_col, "")) if carrier_col else "",
            "actual_weight": parse_number(row.get(weight_col, 0)),
            "actual_quantity": parse_number(row.get(cases_col, 0)),
            "consignee": clean_spaces(row.get(customer_col, "")),
            "source_file": file_name or "MG Report Excel",
        }

        all_records.append(record)

        if load_number not in records_by_load:
            records_by_load[load_number] = record
        else:
            duplicate_mg_loads.append(load_number)

            existing = records_by_load[load_number]

            existing_score = (
                bool(existing.get("consignee")) +
                bool(existing.get("appt_date")) +
                bool(existing.get("appt_time")) +
                bool(existing.get("actual_quantity"))
            )

            new_score = (
                bool(record.get("consignee")) +
                bool(record.get("appt_date")) +
                bool(record.get("appt_time")) +
                bool(record.get("actual_quantity"))
            )

            if new_score > existing_score:
                records_by_load[load_number] = record

    final_records = list(records_by_load.values())
    final_records.sort(key=lambda x: (sort_datetime_key(x), x.get("load", "")))

    summary = {
        "mg_rows_before_dedup": len(all_records),
        "mg_unique_loads": len(final_records),
        "duplicate_mg_loads": sorted(set(duplicate_mg_loads)),
    }

    return final_records, summary


# ============================================================
# OLD MG PDF PARSER KEPT IN CODE, BUT NO LONGER USED FOR MG DATA
# ============================================================

def extract_pdf_pages(pdf_bytes: bytes) -> list[str]:
    pages = []

    with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
        for page in doc:
            pages.append(page.get_text("text"))

    return pages


def extract_mg_load_number(text: str) -> str:
    match = re.search(r"\bLOAD:\s*(?:LD)?\s*([A-Z0-9]+)", text, re.IGNORECASE)

    if not match:
        return ""

    return normalize_load_number(match.group(1))


def extract_mg_pu_appt(text: str) -> tuple[str, str]:
    match = re.search(
        r"\bPU\s+APPT:\s*(\d{1,2}/\d{1,2}/\d{4})\s+(\d{1,2}:\d{2})",
        text,
        re.IGNORECASE,
    )

    if not match:
        return "", ""

    return normalize_date(match.group(1)), normalize_time(match.group(2))


def extract_mg_carrier(text: str) -> str:
    match = re.search(r"\bCARR/SCT\s+TR:\s*(.+)", text, re.IGNORECASE)

    if not match:
        return ""

    carrier = clean_spaces(match.group(1).splitlines()[0])

    if carrier.upper().startswith("CPUC"):
        return "CPUC"

    carrier = re.sub(r"^[A-Z0-9]{3,6}\s+", "", carrier).strip()

    return clean_spaces(carrier)


def extract_mg_totals(text: str) -> tuple[float, float]:
    matches = re.findall(
        r"OUTBOUND\s+TOTALS:\s+([\d,]+(?:\.\d+)?)\s+([\d,]+(?:\.\d+)?)\s+([\d,]+(?:\.\d+)?)",
        text,
        re.IGNORECASE,
    )

    if not matches:
        return 0, 0

    weight, _volume, quantity = matches[-1]

    return parse_number(weight), parse_number(quantity)


def clean_customer_name(value: str) -> str:
    value = clean_spaces(value)

    if not value:
        return ""

    value = re.sub(r"^\d{5,12}\s*/?\s*", "", value)
    value = re.sub(r"^\d{5,12}\s+", "", value)
    value = re.sub(r"\s+TK$", "", value, flags=re.IGNORECASE)
    value = re.sub(r"\s+T$", "", value, flags=re.IGNORECASE)

    bad_fragments = [
        "RESER'S - TK - TOPEKA DISTRIBUTION CENTER",
        "TOPEKA DISTRIBUTION CENTER",
        "3121 SE 6TH AVE",
        "TOPEKA, KS",
        "PH:",
        "FX:",
        "PICKUP TOTAL",
        "DROP TOTAL",
        "OUTBOUND TOTALS",
        "TOTAL MILES",
        "PAGE ",
        "ORDER#",
        "WEIGHT",
        "CUSTID",
        "CUST",
        "NAME",
        "LOCATION",
        "DATE",
        "TIME",
    ]

    upper_value = value.upper()

    for bad in bad_fragments:
        if bad in upper_value:
            return ""

    value = re.sub(r"/\s*\d+[A-Z0-9\-\.]*$", "", value)
    value = re.sub(r"^/\s*", "", value)
    value = clean_spaces(value)

    if re.fullmatch(r"[\d\.\,\-]+", value):
        return ""

    if len(re.sub(r"[^A-Za-z]", "", value)) < 2:
        return ""

    return value


def extract_customers_from_loading_manifest(text: str) -> list[str]:
    customers = []

    for raw_line in text.splitlines():
        line = clean_spaces(raw_line)

        match = re.match(r"^[2-9]\s+(.+)$", line)

        if not match:
            continue

        candidate = clean_customer_name(match.group(1))

        if not candidate:
            continue

        if candidate not in customers:
            customers.append(candidate)

    return customers


def get_shipping_pickup_section(text: str) -> str:
    if "SHIPPING MANIFEST" not in text.upper():
        return ""

    if "Pickup TOTAL" in text:
        return text.split("Pickup TOTAL", 1)[0]

    if "PICKUP TOTAL" in text.upper():
        return re.split(r"PICKUP\s+TOTAL", text, flags=re.IGNORECASE)[0]

    return text


def flush_customer_parts(parts, customers):
    cleaned_parts = []

    for part in parts:
        part = clean_customer_name(part)

        if not part:
            continue

        if part.startswith("/"):
            continue

        cleaned_parts.append(part)

    if not cleaned_parts:
        return

    customer = clean_spaces(" ".join(cleaned_parts))

    customer = customer.replace("SPRINGFIEL D", "SPRINGFIELD")
    customer = customer.replace("GWILLIMBUR Y", "GWILLIMBURY")
    customer = customer.replace("DISTRIBUTO RS", "DISTRIBUTORS")
    customer = customer.replace("PENNSYLVAN IA", "PENNSYLVANIA")
    customer = customer.replace("CONNECTICU T", "CONNECTICUT")
    customer = customer.replace("SOWESTERN ONT", "SW ONTARIO")
    customer = customer.replace("SOUTHWESTERN ONT", "SW ONTARIO")
    customer = clean_spaces(customer)

    if customer and customer not in customers:
        customers.append(customer)


def extract_customers_from_shipping_manifest(text: str) -> list[str]:
    pickup_text = get_shipping_pickup_section(text)

    if not pickup_text:
        return []

    lines = [
        clean_spaces(line)
        for line in pickup_text.splitlines()
        if clean_spaces(line)
    ]

    customers = []
    current_parts = []
    in_order = False

    stop_words = [
        "RESERS FINE FOOD",
        "RESER'S - TK",
        "SHIPPING MANIFEST",
        "PRINTED BY",
        "LOAD:",
        "PU APPT",
        "CARR/SCT",
        "DRIVER",
        "TRACTOR",
        "PRO:",
        "LOCATION",
        "ORDER#",
        "WEIGHT",
        "VOL",
        "PCS",
        "PLT",
        "CUSTID",
        "CUST",
        "NAME",
        "TOPEKA",
        "3121 SE",
        "PH:",
        "APPT:",
        "PICKUP TOTAL",
    ]

    for line in lines:
        upper_line = line.upper()

        if re.match(r"^\d{8,12}-\d{3}", line):
            if current_parts:
                flush_customer_parts(current_parts, customers)

            current_parts = []
            in_order = True

            slash_match = re.search(r"\b[A-Z0-9]{1,10}/\s*(.+)$", line)

            if slash_match:
                possible = clean_customer_name(slash_match.group(1))
                if possible:
                    current_parts.append(possible)

            continue

        if not in_order:
            continue

        if any(word in upper_line for word in stop_words):
            if current_parts:
                flush_customer_parts(current_parts, customers)

            current_parts = []
            in_order = False
            continue

        if line.startswith("/"):
            if current_parts:
                flush_customer_parts(current_parts, customers)

            current_parts = []
            in_order = False
            continue

        if re.fullmatch(r"[\d\.\,\-]+", line):
            continue

        if re.fullmatch(r"\d{2,12}/?", line):
            continue

        if re.match(r"^\d{2,12}/", line):
            after_slash = line.split("/", 1)[1].strip()
            possible = clean_customer_name(after_slash)

            if possible:
                current_parts.append(possible)

            continue

        if re.fullmatch(r"/?[A-Z0-9\-\.\# ]{3,25}", line) and not re.search(r"[AEIOUaeiou]", line):
            continue

        possible = clean_customer_name(line)

        if possible:
            current_parts.append(possible)

    if current_parts:
        flush_customer_parts(current_parts, customers)

    final_customers = []

    for customer in customers:
        customer = clean_spaces(customer)

        if not customer:
            continue

        if len(customer) > 70:
            words = customer.split()
            customer = " ".join(words[:8])

        if customer not in final_customers:
            final_customers.append(customer)

    return final_customers


def is_internal_transfer_customer(customer: str) -> bool:
    upper_customer = customer.upper()

    internal_terms = [
        "RESER'S -",
        "RESERS -",
        "DISTRIBUTION CENTER",
        "CENTURY DISTRIBUTION",
        "HALIFAX DISTRIBUTION",
        "REED TRUCKING",
        "BOZEL TRANSFER",
        "REET -",
        "BOZL -",
        "NC RESER",
        "DC RESER",
    ]

    return any(term in upper_customer for term in internal_terms)


def extract_mg_customers(combined_text: str, loading_text: str, shipping_text: str) -> str:
    shipping_customers = extract_customers_from_shipping_manifest(shipping_text)
    loading_customers = extract_customers_from_loading_manifest(loading_text)

    if shipping_customers:
        selected = shipping_customers
    else:
        selected = loading_customers

    useful_loading = [
        c for c in loading_customers
        if not is_internal_transfer_customer(c)
    ]

    if useful_loading and not shipping_customers:
        selected = useful_loading

    final_customers = []

    for customer in selected:
        customer = clean_customer_name(customer)

        if not customer:
            continue

        if customer not in final_customers:
            final_customers.append(customer)

    return " / ".join(final_customers)


def parse_single_mg_pdf(pdf_bytes: bytes, source_name: str = "") -> list[dict]:
    pages = extract_pdf_pages(pdf_bytes)
    pages_by_load = defaultdict(list)

    for page_text in pages:
        load_number = extract_mg_load_number(page_text)

        if load_number:
            pages_by_load[load_number].append(page_text)

    records = []

    for load_number, load_pages in pages_by_load.items():
        loading_pages = [
            p for p in load_pages
            if "LOADING MANIFEST" in p.upper()
        ]

        shipping_pages = [
            p for p in load_pages
            if "SHIPPING MANIFEST" in p.upper()
        ]

        if loading_pages:
            loading_text = "\n".join(loading_pages)
        else:
            loading_text = "\n".join(load_pages)

        shipping_text = "\n".join(shipping_pages)
        combined_text = loading_text + "\n" + shipping_text

        appt_date, appt_time = extract_mg_pu_appt(loading_text)
        carrier = extract_mg_carrier(loading_text)
        actual_weight, actual_quantity = extract_mg_totals(loading_text)
        consignee = extract_mg_customers(combined_text, loading_text, shipping_text)

        records.append({
            "load": normalize_load_number(load_number),
            "appt_date": appt_date,
            "appt_time": appt_time,
            "carrier": carrier,
            "actual_weight": actual_weight,
            "actual_quantity": actual_quantity,
            "consignee": consignee,
            "source_file": source_name,
        })

    return records


def parse_mg_pdf_bytes(pdf_bytes: bytes, source_name: str = "") -> tuple[list[dict], dict]:
    all_records = parse_single_mg_pdf(pdf_bytes, source_name)
    duplicate_mg_loads = []
    records_by_load = {}

    for record in all_records:
        load = normalize_load_number(record.get("load"))

        if not load:
            continue

        if load not in records_by_load:
            records_by_load[load] = record
        else:
            duplicate_mg_loads.append(load)

            existing = records_by_load[load]

            existing_score = (
                bool(existing.get("consignee")) +
                bool(existing.get("appt_date")) +
                bool(existing.get("appt_time")) +
                bool(existing.get("actual_quantity"))
            )

            new_score = (
                bool(record.get("consignee")) +
                bool(record.get("appt_date")) +
                bool(record.get("appt_time")) +
                bool(record.get("actual_quantity"))
            )

            if new_score > existing_score:
                records_by_load[load] = record

    final_records = list(records_by_load.values())
    final_records.sort(key=lambda x: (sort_datetime_key(x), x.get("load", "")))

    summary = {
        "mg_rows_before_dedup": len(all_records),
        "mg_unique_loads": len(final_records),
        "duplicate_mg_loads": sorted(set(duplicate_mg_loads)),
    }

    return final_records, summary


def parse_multiple_mg_pdfs(uploaded_files) -> tuple[list[dict], dict]:
    all_records = []
    duplicate_mg_loads = []

    for uploaded_file in uploaded_files:
        file_bytes = uploaded_file.read()
        source_name = uploaded_file.name
        records = parse_single_mg_pdf(file_bytes, source_name)
        all_records.extend(records)

    records_by_load = {}

    for record in all_records:
        load = normalize_load_number(record.get("load"))

        if not load:
            continue

        if load not in records_by_load:
            records_by_load[load] = record
        else:
            duplicate_mg_loads.append(load)

            existing = records_by_load[load]

            existing_score = (
                bool(existing.get("consignee")) +
                bool(existing.get("appt_date")) +
                bool(existing.get("appt_time")) +
                bool(existing.get("actual_quantity"))
            )

            new_score = (
                bool(record.get("consignee")) +
                bool(record.get("appt_date")) +
                bool(record.get("appt_time")) +
                bool(record.get("actual_quantity"))
            )

            if new_score > existing_score:
                records_by_load[load] = record

    final_records = list(records_by_load.values())
    final_records.sort(key=lambda x: (sort_datetime_key(x), x.get("load", "")))

    summary = {
        "mg_rows_before_dedup": len(all_records),
        "mg_unique_loads": len(final_records),
        "duplicate_mg_loads": sorted(set(duplicate_mg_loads)),
    }

    return final_records, summary


# ============================================================
# TEMPLATE POPULATION
# ============================================================

def validate_template(wb):
    missing = []

    for sheet_name in TEMPLATE_REQUIRED_SHEETS:
        if sheet_name not in wb.sheetnames:
            missing.append(sheet_name)

    if missing:
        raise ValueError("Template is missing these sheets: " + ", ".join(missing))


def populate_opendock_sheet(wb, opendock_records):
    ws = wb[OPENDOCK_SHEET]

    max_row = max(ws.max_row, OPENDOCK_START_ROW + MAX_OPENDOCK_ROWS)

    clear_range_values(ws, OPENDOCK_START_ROW, max_row, [1, 2, 3, 4, 5, 6, 7])

    for index, record in enumerate(opendock_records, start=OPENDOCK_START_ROW):
        copy_row_format(ws, OPENDOCK_START_ROW, index, 7)

        ws.cell(index, 1).value = record.get("load", "")
        ws.cell(index, 2).value = record.get("arrival_date", "")
        ws.cell(index, 3).value = record.get("arrival_time", "")
        ws.cell(index, 4).value = record.get("departure_date", "")
        ws.cell(index, 5).value = record.get("departure_time", "")
        ws.cell(index, 6).value = record.get("status", "")
        ws.cell(index, 7).value = record.get("carrier", "")


def populate_mg_report_sheet(wb, mg_records):
    ws = wb[MG_REPORT_SHEET]

    max_row = max(ws.max_row, MG_REPORT_START_ROW + MAX_MG_REPORT_ROWS)

    clear_range_values(ws, MG_REPORT_START_ROW, max_row, [1, 2, 3, 4])

    for index, record in enumerate(mg_records, start=MG_REPORT_START_ROW):
        copy_row_format(ws, MG_REPORT_START_ROW, index, 4)

        ws.cell(index, 1).value = record.get("load", "")
        ws.cell(index, 2).value = record.get("actual_weight", 0)
        ws.cell(index, 3).value = record.get("actual_quantity", 0)
        ws.cell(index, 4).value = record.get("consignee", "")


def populate_dispatch_sheet(wb, opendock_records, mg_records):
    ws = wb[DISPATCH_SHEET]

    cols = get_dispatch_columns(ws, header_row=2)
    notes_col = cols["notes"]

    # Letters used inside the Excel formulas (resolved from headers)
    col_load = get_column_letter(cols["load"])
    col_customer = get_column_letter(cols["customer"])
    col_type = get_column_letter(cols["type"])
    col_carrier_full = get_column_letter(cols["carrier_full"])

    max_output_col = max(ws.max_column, *cols.values())

    row_template = capture_row_template(
        ws,
        DISPATCH_START_ROW,
        max_output_col
    )

    mg_by_load = {
        normalize_load_number(record.get("load")): record
        for record in mg_records
        if normalize_load_number(record.get("load"))
    }

    max_row = max(ws.max_row, DISPATCH_START_ROW + MAX_DISPATCH_ROWS)

    clear_dispatch_area(
        ws,
        DISPATCH_START_ROW,
        max_row,
        max_output_col
    )

    grouped_by_date = []

    for record in opendock_records:
        date_text = record.get("appt_date", "") or record.get("arrival_date", "") or "NO DATE"

        if not grouped_by_date or grouped_by_date[-1]["date"] != date_text:
            grouped_by_date.append({
                "date": date_text,
                "records": []
            })

        grouped_by_date[-1]["records"].append(record)

    current_row = DISPATCH_START_ROW
    loads_written = 0

    for day_group in grouped_by_date:
        date_text = day_group["date"]
        day_records = day_group["records"]

        day_cases = 0

        for record in day_records:
            load_number = normalize_load_number(record.get("load"))
            matched_mg = mg_by_load.get(load_number)

            if matched_mg:
                day_cases += parse_number(matched_mg.get("actual_quantity", 0))

        if current_row > DISPATCH_START_ROW + MAX_DISPATCH_ROWS:
            break

        apply_row_template(
            ws,
            row_template,
            current_row,
            copy_values=False
        )

        write_day_header(
            ws,
            current_row,
            date_text,
            len(day_records),
            day_cases,
            max_output_col
        )

        current_row += 1

        for record in day_records:
            if current_row > DISPATCH_START_ROW + MAX_DISPATCH_ROWS:
                break

            apply_row_template(
                ws,
                row_template,
                current_row,
                copy_values=True
            )

            load_number = record.get("load", "")

            if not load_number:
                current_row += 1
                continue

            matched_mg = mg_by_load.get(load_number)

            notes = []

            if not matched_mg:
                notes.append("Not found in MG report")
            else:
                mg_date = matched_mg.get("appt_date", "")
                mg_time = matched_mg.get("appt_time", "")
                od_date = record.get("appt_date", "")
                od_time = record.get("appt_time", "")

                if mg_date and od_date and mg_date != od_date:
                    notes.append(f"Date mismatch: MG {mg_date}, Opendock {od_date}")

                if mg_time and od_time and mg_time != od_time:
                    notes.append(f"Time mismatch: MG {mg_time}, Opendock {od_time}")

                if not matched_mg.get("consignee", ""):
                    notes.append("MG customer/consignee not found")

            row_num = current_row

            # --- Directly written values ---
            ws.cell(row_num, cols["load"]).value = load_number
            ws.cell(row_num, cols["time"]).value = record.get("appt_time", "")
            ws.cell(row_num, cols["carrier_full"]).value = record.get("carrier", "")
            ws.cell(row_num, cols["type"]).value = clean_load_type(record.get("load_type", ""))
            ws.cell(row_num, notes_col).value = " | ".join(notes)

            # --- Formula-driven cells (column letters resolved from headers) ---
            # CUSTOMER: consignee from MG REPORT col 4
            ws.cell(row_num, cols["customer"]).value = (
                f'=IFERROR(VLOOKUP({col_load}{row_num},\'MG REPORT\'!$A$2:$D$1001,4,FALSE),"")'
            )
            # CARRIER: short code from CARRIERS based on CARRIER FULL.
            # If no carrier is found (blank) AND the load is a CPU -- i.e. the
            # CARRIER FULL text or the TYPE says "CPU" / "Customer Pickup"
            # -- show "CPU"; otherwise leave blank.
            ws.cell(row_num, cols["carrier"]).value = (
                f'=IFERROR(VLOOKUP({col_carrier_full}{row_num},CARRIERS!$A:$B,2,FALSE),'
                f'IF(OR('
                f'ISNUMBER(SEARCH("CPU",{col_carrier_full}{row_num})),'
                f'ISNUMBER(SEARCH("CUSTOMER PICK",{col_carrier_full}{row_num})),'
                f'ISNUMBER(SEARCH("CPU",{col_type}{row_num})),'
                f'ISNUMBER(SEARCH("CUSTOMER PICK",{col_type}{row_num}))),'
                f'"CPU",""))'
            )
            # TT4: flag based on CUSTOMER
            ws.cell(row_num, cols["tt4"]).value = (
                f'=IF(OR(ISNUMBER(SEARCH("albertson",{col_customer}{row_num})),'
                f'ISNUMBER(SEARCH("jewel",{col_customer}{row_num})),'
                f'ISNUMBER(SEARCH("safeway",{col_customer}{row_num})),'
                f'ISNUMBER(SEARCH("sysco",{col_customer}{row_num})),'
                f'ISNUMBER(SEARCH("united supermarkets",{col_customer}{row_num}))),"X","")'
            )
            # WEIGHT: MG REPORT col 2
            ws.cell(row_num, cols["weight"]).value = (
                f'=IFERROR(VALUE(VLOOKUP({col_load}{row_num},\'MG REPORT\'!$A$2:$D$1001,2,FALSE)),0)'
            )
            # CASES: MG REPORT col 3
            ws.cell(row_num, cols["cases"]).value = (
                f'=IFERROR(VALUE(VLOOKUP({col_load}{row_num},\'MG REPORT\'!$A$2:$D$1001,3,FALSE)),0)'
            )

            # PULLS / PICKS keep their template formulas; just fix number format
            if cols.get("pulls"):
                ws.cell(row_num, cols["pulls"]).number_format = "0"
            if cols.get("picks"):
                ws.cell(row_num, cols["picks"]).number_format = "0"

            loads_written += 1
            current_row += 1

    count_cell = find_dispatch_count_cell(ws)
    if count_cell:
        ws[count_cell] = loads_written

    date_cell = find_dispatch_date_cell(ws)
    if date_cell:
        if len(grouped_by_date) == 1:
            ws[date_cell] = grouped_by_date[0]["date"]
        elif len(grouped_by_date) > 1:
            ws[date_cell] = "Multiple Dates"
        else:
            ws[date_cell] = ""


def add_match_report_sheet(wb, opendock_records, mg_records):
    sheet_name = "MATCH REPORT"

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    headers = [
        "Load",
        "Opendock Date",
        "Opendock Time",
        "MG Date",
        "MG Time",
        "Opendock Carrier",
        "MG Carrier",
        "Opendock Load Type",
        "MG Customer(s)",
        "MG Weight",
        "MG Cases",
        "MG Source File",
        "Status",
    ]

    ws.append(headers)

    mg_by_load = {
        normalize_load_number(record.get("load")): record
        for record in mg_records
        if normalize_load_number(record.get("load"))
    }

    opendock_loads = set()

    for record in opendock_records:
        load_number = record.get("load", "")
        opendock_loads.add(load_number)
        mg = mg_by_load.get(load_number)

        if not mg:
            status = "Missing in MG report"
            mg_date = ""
            mg_time = ""
            mg_carrier = ""
            consignee = ""
            source_file = ""
            weight = 0
            cases = 0
        else:
            mg_date = mg.get("appt_date", "")
            mg_time = mg.get("appt_time", "")
            mg_carrier = mg.get("carrier", "")
            consignee = mg.get("consignee", "")
            weight = mg.get("actual_weight", 0)
            cases = mg.get("actual_quantity", 0)
            source_file = mg.get("source_file", "")

            status_parts = []

            if mg_date and record.get("appt_date") and mg_date != record.get("appt_date"):
                status_parts.append("Date mismatch")

            if mg_time and record.get("appt_time") and mg_time != record.get("appt_time"):
                status_parts.append("Time mismatch")

            if not consignee:
                status_parts.append("Customer missing from MG")

            if status_parts:
                status = " / ".join(status_parts)
            else:
                status = "Matched"

        ws.append([
            load_number,
            record.get("appt_date", ""),
            record.get("appt_time", ""),
            mg_date,
            mg_time,
            record.get("carrier", ""),
            mg_carrier,
            record.get("load_type", ""),
            consignee,
            weight,
            cases,
            source_file,
            status,
        ])

    for mg_load, mg in mg_by_load.items():
        if mg_load not in opendock_loads:
            ws.append([
                mg_load,
                "",
                "",
                mg.get("appt_date", ""),
                mg.get("appt_time", ""),
                "",
                mg.get("carrier", ""),
                "",
                mg.get("consignee", ""),
                mg.get("actual_weight", 0),
                mg.get("actual_quantity", 0),
                mg.get("source_file", ""),
                "Missing in Opendock outbound report",
            ])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for column in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(column)].width = 20

    ws.column_dimensions["I"].width = 55
    ws.column_dimensions["L"].width = 35
    ws.column_dimensions["M"].width = 35

    ws.freeze_panes = "A2"


def format_output_workbook(wb):
    for sheet_name in [DISPATCH_SHEET, OPENDOCK_SHEET, MG_REPORT_SHEET]:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        # Resolve the wide columns by header so widths follow the layout.
        dispatch_customer_letter = None
        dispatch_notes_letter = None
        if sheet_name == DISPATCH_SHEET:
            header_map = get_header_map(ws, 2)
            cust_col = find_col(header_map, DISPATCH_HEADER_ALIASES["customer"])
            note_col = find_col(header_map, DISPATCH_HEADER_ALIASES["notes"])
            if cust_col:
                dispatch_customer_letter = get_column_letter(cust_col)
            if note_col:
                dispatch_notes_letter = get_column_letter(note_col)

        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)

            if sheet_name == DISPATCH_SHEET:
                # Leave hidden columns alone so they stay hidden/sized.
                if ws.column_dimensions[letter].hidden:
                    continue

                if letter == dispatch_customer_letter:      # CUSTOMER
                    ws.column_dimensions[letter].width = 32
                elif letter == dispatch_notes_letter:        # NOTES
                    ws.column_dimensions[letter].width = 50
                else:
                    ws.column_dimensions[letter].width = 14

            elif sheet_name == OPENDOCK_SHEET:
                ws.column_dimensions[letter].width = 18

            elif sheet_name == MG_REPORT_SHEET:
                if letter == "D":
                    ws.column_dimensions[letter].width = 60
                else:
                    ws.column_dimensions[letter].width = 18

        try:
            ws.freeze_panes = "A2"
        except Exception:
            pass


def populate_template(template_bytes: bytes, opendock_records: list[dict], mg_records: list[dict]) -> bytes:
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))

    validate_template(wb)
    populate_opendock_sheet(wb, opendock_records)
    populate_mg_report_sheet(wb, mg_records)
    populate_dispatch_sheet(wb, opendock_records, mg_records)
    add_match_report_sheet(wb, opendock_records, mg_records)
    format_output_workbook(wb)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


# ============================================================
# STREAMLIT APP
# ============================================================

st.set_page_config(
    page_title="Dispatch Builder",
    layout="wide"
)

st.title("Dispatch Builder")

st.write(
    "Upload the Loading Manifest PDF, Shipping Manifest PDF, "
    "the Opendock report, and the MG Report Excel. The dispatch template is read from "
    "the Dispatch_Template.xlsx file next to this app. "
    "The app will match and merge the two manifests, "
    "then populate the OPENDOCK, MG REPORT, DISPATCH SHEET, and MATCH REPORT tabs. "
    "If the Opendock report has multiple appointment dates, the Dispatch Sheet is separated by date."
)

st.subheader("Step 1 — Upload Manifests")

col_left, col_right = st.columns(2)

with col_left:
    loading_manifest_file = st.file_uploader(
        "Loading Manifest — PDF or ZIP",
        type=["pdf", "zip"],
        key="loading_manifest"
    )

with col_right:
    shipping_manifest_file = st.file_uploader(
        "Shipping Manifest — PDF or ZIP",
        type=["pdf", "zip"],
        key="shipping_manifest"
    )

st.subheader("Step 2 — Upload Opendock Report")

st.caption("The dispatch template is read from Dispatch_Template.xlsx next to this app — edit that file in Excel to change it.")

opendock_file = st.file_uploader(
    "Opendock Report (.xlsx)",
    type=["xlsx"],
    key="opendock"
)

st.subheader("Step 3 — Upload MG Report Excel")

mg_report_file = st.file_uploader(
    "MG Report Excel (.xlsx or .xls)",
    type=["xlsx", "xls"],
    key="mg_report"
)

st.divider()

all_required = (
    loading_manifest_file
    and shipping_manifest_file
    and opendock_file
    and mg_report_file
)

if not all_required:
    st.info("Upload the Loading Manifest, Shipping Manifest, Opendock report, and MG Report Excel to continue.")

if all_required and st.button("Build Matched PDF + Populated Short Sheet", type="primary"):

    try:
        loading_bytes = extract_pdf_bytes_from_upload(loading_manifest_file)
        shipping_bytes = extract_pdf_bytes_from_upload(shipping_manifest_file)
        template_bytes = get_embedded_template_bytes()
        opendock_bytes = opendock_file.read()
        mg_report_bytes = mg_report_file.read()

        with st.spinner("Matching and merging manifest PDFs…"):
            matched_pdf_bytes, loading_count, shipping_count, load_count = build_matched_pdf_bytes(
                loading_bytes,
                shipping_bytes
            )

        st.success(
            f"Matched PDF built — {loading_count} loading loads, "
            f"{shipping_count} shipping loads, {load_count} unique loads."
        )

        st.download_button(
            label="Download Matched Manifest PDF",
            data=matched_pdf_bytes,
            file_name="Matched_Manifest_Packet.pdf",
            mime="application/pdf",
        )

        with st.spinner("Reading MG Report Excel…"):
            mg_records, mg_summary = parse_mg_report_excel(
                mg_report_bytes,
                mg_report_file.name
            )

        with st.spinner("Reading Opendock report…"):
            opendock_records, opendock_summary = parse_opendock_excel(opendock_bytes)

        st.success("All files loaded successfully.")

        c1, c2, c3, c4 = st.columns(4)

        c1.metric("Opendock Outbound Rows", opendock_summary["outbound_rows_loaded"])
        c2.metric("Inbound Rows Skipped", opendock_summary["inbound_rows_skipped"])
        c3.metric("MG Unique Loads", mg_summary["mg_unique_loads"])
        c4.metric("MG Rows Before Dedup", mg_summary["mg_rows_before_dedup"])

        if opendock_summary["duplicate_outbound_loads"]:
            st.warning(
                "Duplicate outbound loads in Opendock: "
                + ", ".join(opendock_summary["duplicate_outbound_loads"])
            )

        if mg_summary["duplicate_mg_loads"]:
            st.warning(
                "Duplicate loads found in MG Report Excel — kept the most complete version: "
                + ", ".join(mg_summary["duplicate_mg_loads"])
            )

        with st.spinner("Populating template…"):
            populated_file = populate_template(
                template_bytes,
                opendock_records,
                mg_records
            )

        st.success("Populated short sheet created successfully.")

        st.subheader("Opendock Outbound Preview")
        st.dataframe(opendock_records, use_container_width=True)

        st.subheader("MG Report Preview")
        st.dataframe(mg_records, use_container_width=True)

        missing_customer_records = [
            r for r in mg_records
            if not r.get("consignee", "")
        ]

        if missing_customer_records:
            st.warning(f"{len(missing_customer_records)} MG load(s) have no customer/consignee detected.")
            st.dataframe(missing_customer_records, use_container_width=True)
        else:
            st.success("Customer/consignee populated for all MG loads.")

        st.download_button(
            label="Download Populated Dispatch Template",
            data=populated_file,
            file_name=OUTPUT_FILE_NAME,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # ── Download matched manifest — always visible after build ──────────

        st.download_button(

            label=" Download Matched Manifest PDF",

            data=matched_pdf_bytes,

            file_name="Matched_Manifest_Packet.pdf",

            mime="application/pdf",

            key="download_matched_pdf",

        )

    except Exception as e:
        st.error("Something went wrong.")
        st.exception(e)
